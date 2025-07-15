import pandas as pd
from openpyxl import load_workbook

from function.dowload_to_pandas import dowload_df_filename

def map_excel_columns(
    source_file: str,
    destination_file: str,
    source_sheet: str,
    destination_sheet: str,
    column_mapping: dict,
    source_header_row: int = 0
) -> pd.DataFrame:
    """
    Maps and transfers data from a source Excel file to a destination Excel file template.
    Supports both one-to-one and one-to-many column mappings.
    """

    try:
        print(f"📖 Reading source file: {source_file} (Sheet: {source_sheet})")

        # # Use openpyxl to read actual headers from source
        # wb = load_workbook(source_file, data_only=True)
        # ws = wb[source_sheet]
        # header_row_cells = ws[source_header_row + 1]  # openpyxl is 1-based
        # source_headers = [cell.value if cell.value is not None else f"Unnamed: {i}" 
        #                   for i, cell in enumerate(header_row_cells)]

        # # Read source DataFrame
        # source_df = pd.read_excel(
        #     source_file, 
        #     sheet_name=source_sheet, 
        #     header=None, 
        #     skiprows=source_header_row + 1,  
        #     engine='openpyxl'
        # )
        # source_df.columns = source_headers
        # source_df.columns = source_df.columns.astype(str).str.strip()

        # print(f"📖 Reading destination template headers from: {destination_file} (Sheet: {destination_sheet})")
        source_df = dowload_df_filename(source_file, source_sheet)
        
        destination_df = pd.read_excel(
            destination_file, 
            sheet_name=destination_sheet, 
            nrows=0,  # read only headers
            engine='openpyxl'
        )
        destination_columns = destination_df.columns.astype(str).str.strip()
        # Create new DataFrame for mapped data with destination template headers
        new_df = pd.DataFrame(index=source_df.index, columns=destination_columns)

        for src_col, dest_cols in column_mapping.items():
            # Allow dest_cols to be a list or a single value
            if not isinstance(dest_cols, list):
                dest_cols = [dest_cols]
        

            if src_col in source_df.columns:
                for dest_col in dest_cols:
                    if dest_col in new_df.columns:
                        new_df[dest_col] = source_df[src_col].values
                        print(f"✅ Mapped '{src_col}' → '{dest_col}'")
                    else:
                        print(f"⚠️ Destination column '{dest_col}' not found in template.")
            else:
                print(f"⚠️ Source column '{src_col}' not found in source file.")

        print(f"\n🎉 Mapping complete! Ready to export.")
        return new_df

    except Exception as e:
        print(f"❌ Error occurred: {str(e)}")
        raise e


# Example usage
# source_file = r"C:\Users\andre\Desktop\Internship Data cleaning\ข้อมูลคนค้ำ+ที่อยู่คนค้ำ WL\wl_zqcredit_04.2025.xlsx"
# destination_file = r"C:\Users\andre\Desktop\Internship Data cleaning\ข้อมูลคนค้ำ+ที่อยู่คนค้ำ WL\Template 1.3 - GRT Testing.xlsx"
# source_sheet = "Sheet1"
# destination_sheet = "Template1"

# column_mapping = {
#     'รหัส': 'รหัสลูกค้า',
#     'เลขที่บัตรประชาชน': 'เลขที่บัตร',
#     'โทรศัพท์': 'โทรศัพท์มือถือ'
# }

# # Map data
# mapped_df = map_excel_columns(
#     source_file,
#     destination_file,
#     source_sheet,
#     destination_sheet,
#     column_mapping
# # )