from openpyxl import load_workbook
from datetime import datetime

def define_customer_type(source_file, workbook, source_column, target_column):
    # Save output with timestamp to avoid overwrite 
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    # Ensure path is absolute and normalized
    output_file = source_file

    # Load source workbook and worksheet
    source_wb = load_workbook(source_file)
    source_ws = source_wb[workbook]
    source_header = [cell.value for cell in source_ws[1]]

    # Array of company's prefix
    prefix_company = [
        "บริษัท",
        "บจก.",
        "ห้างหุ้นส่วนจำกัด",
        "หจก.",
        "ห้างหุ้นส่วน",
        "บมจ.",
        "มหาชน",
        "ร้าน"
    ]

    # Validate column names
    if source_column not in source_header:
        print("⚠️ Source column \"" + source_column + "\" not found. Skipping.")
        return
    else:
        print("✅ Found source column: " + source_column)

    if target_column not in source_header:
        print("⚠️ Target column \"" + target_column + "\" not found. Skipping.")
        return
    else:
        print("✅ Found target column: " + target_column)

    # Get column indexes (1-indexed)
    source_col_index = source_header.index(source_column) + 1
    target_col_index = source_header.index(target_column) + 1

    # Loop through each row starting from row 2
    for row_idx in range(2, source_ws.max_row + 1):
        value = ""
        cell_value = source_ws.cell(row=row_idx, column=source_col_index).value

        if cell_value in prefix_company:
            value = "C"
        elif cell_value is None or cell_value == "":
            value = ""
        else:
            value = "I"

        # Write result
        source_ws.cell(row=row_idx, column=target_col_index, value=value)

    print(f'✅ Define type of customer from "{source_column}" to "{target_column}".')

    # Save file
    source_wb.save(output_file)
    print(f'🎉 Done! Updated file saved as:\n{output_file}')