from openpyxl import load_workbook

def update_xlsx_with_clean_names_openpyxl(file_path, cleaned_data, column_map, sheet_name='Template1'):
    """
    file_path: path to Excel file (.xlsx)
    cleaned_data: list of dicts returned by clean_name(), each with cleaned fields
    column_map: dict mapping cleaned field name to Excel header
                e.g. { 'คำนำหน้า': 'Prefix', 'ชื่อ': 'First Name', 'นามสกุล': 'Last Name' }
    sheet_name: name of sheet to update
    """

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in file '{file_path}'")
    ws = wb[sheet_name]

    # Map headers to column indices
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_to_index = {header: idx + 1 for idx, header in enumerate(headers)}

    # Ensure all destination columns exist or add them to the header row
    max_col = len(headers)
    for clean_col, excel_col in column_map.items():
        if excel_col not in header_to_index:
            max_col += 1
            header_to_index[excel_col] = max_col
            ws.cell(row=1, column=max_col).value = excel_col
            print(f"⚠️ Created missing column '{excel_col}'")

    # Update each row with cleaned data
    for i, row_data in enumerate(cleaned_data, start=2):  # Start from row 2 (after header)
        for clean_col, excel_col in column_map.items():
            value = row_data.get(clean_col, None)
            col_idx = header_to_index[excel_col]
            ws.cell(row=i, column=col_idx).value = value

    # Save workbook
    wb.save(file_path)
    print(f"✅ Updated file: {file_path}")




