from openpyxl import load_workbook
import os

def map_column_WM(source_file: str, destination_file: str) -> str:
    """
    Copies and transforms data from a source Excel file into a destination template (Template 1.2 or 2.2),
    fills in default values, and overwrites the destination file.

    Parameters:
        source_file (str): Path to the source Excel file.
        destination_file (str): Path to the destination Excel template (Template 1.2 or 2.2).

    Returns:
        str: Path to the overwritten destination file.
    """

    # Column mapping between source and both templates
    column_mapping = {
        'ลูกค้า': 'รหัสลูกค้า',
        'เลขที่จดทะเบียน VAT': 'เลขที่บัตร',
        'วันเดือนปีเกิด': 'วันเดือนปีเกิด',
        'อายุ': 'อายุ',
        'อาชีพ': 'อาชีพ/ประเภทธุรกิจ',
        'รายได้(เดือน)': 'รายได้ประจำ',
        'โทรศัพท์': ['โทรศัพท์มือถือ', 'Phone'],  # use list for fallbacks
        '1. ชื่อ-สกุล บุคคลอ้างอิง 1': 'บุคคลอ้างอิง 1',
        'ความสัมพันธ์บุคคลอ้างอิง 1': 'ความสัมพันธ์ 1',
        'เบอร์โทรบุคคลอ้างอิง 1': 'เบอร์โทรบุคคลอ้างอิง 1',
        '2. ชื่อ-สกุล บุคคลอ้างอิง 2': 'บุคคลอ้างอิง 2',
        'ความสัมพันธ์บุคคลอ้างอิง 2': 'ความสัมพันธ์ 2',
        'เบอร์โทรบุคคลอ้างอิง 2': 'เบอร์โทรบุคคลอ้างอิง 2'
    }

    # Load source workbook
    source_wb = load_workbook(source_file, data_only=True)
    source_ws = source_wb.active
    source_headers = [cell.value for cell in source_ws[1]]
    source_col_indices = {col: idx + 1 for idx, col in enumerate(source_headers)}

    # Read source data
    source_data = {col: [] for col in source_headers}
    for row in source_ws.iter_rows(min_row=2, max_row=source_ws.max_row, values_only=True):
        for col_name, col_idx in source_col_indices.items():
            source_data[col_name].append(row[col_idx - 1])

    # Load destination workbook and find target sheet
    wb = load_workbook(destination_file)
    possible_sheets = ['Template1_WM', 'Template2_WM']
    ws = None
    for sheet in possible_sheets:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            break
    if ws is None:
        raise Exception("❌ Could not find a known template sheet in the destination file.")

    # Destination headers
    header = [cell.value for cell in ws[1]]

    # Copy data based on mapping
    for src_col, dest_col in column_mapping.items():
        if src_col not in source_data:
            print(f'⚠️ Source column "{src_col}" not found. Skipping.')
            continue

    # Handle fallback destination (if list of options)
        if isinstance(dest_col, list):
            chosen_dest_col = None
            for col_option in dest_col:
                if col_option in header:
                    chosen_dest_col = col_option
                    break
            if not chosen_dest_col:
                print(f'⚠️ None of destination columns {dest_col} found for "{src_col}". Skipping.')
                continue
        else:
            if dest_col not in header:
                print(f'⚠️ Destination column "{dest_col}" not found. Skipping.')
                continue
            chosen_dest_col = dest_col

        dest_col_index = header.index(chosen_dest_col) + 1
        for row_idx, value in enumerate(source_data[src_col], start=2):
            ws.cell(row=row_idx, column=dest_col_index, value=value)

        print(f'✅ Copied "{src_col}" → "{chosen_dest_col}"')

    # Fill default values
    default_fills = {
        'ประเภทบัตร': '1',
        'ประเภทที่อยู่': '1',
        'ใช้ส่งเอกสาร': 'Y'
    }
    for col_name, default_value in default_fills.items():
        if col_name in header:
            col_index = header.index(col_name) + 1
            for row in ws.iter_rows(min_row=2):
                row[col_index - 1].value = default_value
            print(f'✅ Filled column "{col_name}" with "{default_value}"')
        else:
            print(f'⚠️ Column "{col_name}" not found. Skipped filling default.')

    # Overwrite the destination file directly
    wb.save(destination_file)
    print(f'\n🎉 Done! File updated in place:\n{destination_file}')

    return destination_file