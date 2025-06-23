import pandas as pd
from rapidfuzz import process, fuzz
from Levenshtein import distance as levenshtein_distance
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
from openpyxl import load_workbook
import re

## define variable
# for 1st function
exel_file = "wl_zqcredit_04.2025.xlsx" 
name_column = "ชื่อ-สกุล"

# for 2nd function
issuenote_column = "ผลการตรวจสอบ" # column that assign issue
note_txt = "ตรวจสอบคำนำหน้า" # output text
target_column = "คำนำหน้า" # column name that need to check
target_value = {'ร.ต.อ', 'ส.ต.อ'} # prefix that need to issue note on
# df = df name that is a result of "clean_name" function

## 1st function : use to split prefix, name, surname from full name
def clean_name(source_file, sheet_name, name_column):
    
    original_df = pd.read_excel(source_file, sheet_name=sheet_name, engine="openpyxl")
    # Use openpyxl to read actual headers from source
    
    wb = load_workbook(source_file, data_only=True)
    ws = wb[sheet_name]
    header_row_cells = ws[0 + 1]  # openpyxl is 1-based
    source_headers = [cell.value if cell.value is not None else f"Unnamed: {i}" 
                      for i, cell in enumerate(header_row_cells)]
    # print(source_headers)
    
    original_df.columns = source_headers
    
    
    df = original_df[[name_column]].copy()
    
    # all prefix that we could check
    thai_english_prefixes = [
            # เด็ก
            "เด็กชาย", "ด.ช.","เด็กหญิง", "ด.ญ.","นาวสาว",
            # ผู้ใหญ่ทั่วไป
            "นาย", "นาง", "นางสาว", "น.ส.", "าง","คุณ", "นาน",
            # วิชาการ
            "ดอกเตอร์", "ดร.", "ศาสตราจารย์", "ผู้ช่วยศาสตราจารย์", "ผศ.", "รองศาสตราจารย์", "รศ.", "ศาสตราจารย์ ดร.", "ศ.ดร.", "ผศ.ดร.", "รศ.ดร.",
            # ทหาร / ตำรวจ
            "พลเอก", "พล.อ.", "พลโท", "พล.ท.", "พลตรี", "พล.ต.", "พันเอก", "พ.อ.", "พันโท", "พ.ท.", "พันตรี", "พ.ต.", "พ.ต.ท.", "พ.ต.ท", "พ.ต.อ.", "ร้อยเอก", 
            "ร.อ.", "ร้อยโท", "ร.ท.", "ร้อยตรี", "ร.ต.","ร.ต.หญิง", "ร.ต.ต.", "ร.ต.อ.","ร.ต.อ",  "สิบตรี", "ส.ต.", "พลตำรวจเอก", "พล.ต.อ.", "พันตำรวจเอก", "พ.ต.อ.",
            "จ.ส.อ.", "ด.ต.", "ส.ท.", "จ.ส.ต.", "ว่าที่ ร.ต.หญิง","ส.ณ.", "ส.อ.", "จ.อ.", "ดาบตำรวจ", "พ.อ.อ.", "พ.อ.ต.", "ร.ต.ท.", "ส.ต.อ.",
            "ส.ต.อ","ส.ต.ต.","ส.ต.ท.","พลตำรวจเอก", "พล.ต.อ.", "พันตำรวจเอก","ร.ต.หญิง", 
            # ราชวงศ์ / ขุนนาง
            "หม่อม",  "หม่อมหลวง", "ม.ล.", "หม่อมราชวงศ์", "ม.ร.ว.", "หม่อมเจ้า", "ม.จ.", "คุณหญิง", "คุณชาย", "ท่านผู้หญิง",
            # วิชาชีพ
            "นายแพทย์", "นพ.", "ทันตแพทย์", "ทพ.", "เภสัชกร", "ภก.", "ภกญ.", "สัตวแพทย์", "สพ.", "วิศวกร", "วศ.", "นิติกร", "นก.", "บริษัท", "หจก.", "ห้างหุ้นส่วนจำกัด",
            "พระ", 
            # English Prefix
            "MR.", "Mrs.", "Ms.", "Miss", "Dr.", "Prof.",  "Sir", "Madam", "Lady", "Lord", "Rev.", "Capt.",  "Col.", "Gen.", "Lt.", "Sgt.", "Major"
        ]
        
    
    # ในกรณีที่คำนำหน้ามีคำว่า " ว่าที่ "
    # Add "ว่าที่" versions
    all_prefixes = []
    for p in thai_english_prefixes:
        all_prefixes.append(p)
        if not p.startswith("ว่าที่"):
            all_prefixes.append("ว่าที่" + p)
            all_prefixes.append("ว่าที่ " + p)
            all_prefixes.append("ว่าที" + p)
            all_prefixes.append("ว่าที " + p)



    def extract_parts(full_name, threshold=85):
        if pd.isnull(full_name) or not isinstance(full_name, str):
            return pd.Series([None, None, None])
    
        name_clean = full_name.strip()
        prefix_raw = None
        rest = name_clean
    
        ## 1. Exact Match (with original prefixes)
        # Create a sorted list of original prefixes by length (longest first)
        original_prefixes_sorted = sorted(all_prefixes, key=lambda x: -len(x))
    
        for original_prefix in original_prefixes_sorted:
            if name_clean.startswith(original_prefix):
                prefix_raw = original_prefix
                rest = name_clean[len(original_prefix):].strip()
                break # Exit loop once an exact match is found
    
        ## 2. Normalized Match (if no exact match found)
        if prefix_raw is None:
            name_for_match = re.sub(r"[.\s]", "", name_clean).lower()
    
            # Prepare normalized prefix keys and map
            normalized_prefix_map = {p.replace(".", "").replace(" ", "").lower(): p for p in all_prefixes}
            prefix_keys_sorted = sorted(normalized_prefix_map.keys(), key=lambda x: -len(x))
    
            # 2.1 Exact match (longest normalized prefix first)
            for key in prefix_keys_sorted:
                if name_for_match.startswith(key):
                    prefix_raw = normalized_prefix_map[key]
                    
                    # To get the correct 'rest' for normalized matches, we need to find 
                    # where the matched normalized 'key' ends in the original name_clean.
                    # We'll do this by building a normalized version of name_clean character by character
                    # and finding the corresponding end index in the original name_clean.
                    original_prefix_end_index = 0
                    k_idx = 0
                    for char_idx, char in enumerate(name_clean):
                        cleaned_char = re.sub(r"[.\s]", "", char).lower()
                        if cleaned_char:
                            if k_idx < len(key) and cleaned_char == key[k_idx]:
                                k_idx += 1
                            else:
                                # If we've processed all characters of the 'key'
                                if k_idx == len(key):
                                    original_prefix_end_index = char_idx
                                    break
                    if k_idx == len(key) and original_prefix_end_index == 0:
                        original_prefix_end_index = len(name_clean)
    
                    rest = name_clean[original_prefix_end_index:].strip()
                    break
    
            # 2.2 Autocorrect fallback if no exact normalized match
            if prefix_raw is None:
                snippets = [name_for_match[:i] for i in range(2, min(20, len(name_for_match) + 1))]
    
                best_match = ("", float("inf"))
                for snippet in snippets:
                    for prefix_key in prefix_keys_sorted:
                        dist = levenshtein_distance(snippet, prefix_key)
                        if dist < best_match[1]:
                            best_match = (prefix_key, dist)
    
                match, dist = best_match
                max_distance = 2  # You can adjust this
                if dist <= max_distance and name_for_match.startswith(match):
                    prefix_raw = normalized_prefix_map[match]
                    
                    # Recalculate original_prefix_end_index for autocorrected match
                    original_prefix_end_index = 0
                    k_idx = 0
                    for char_idx, char in enumerate(name_clean):
                        cleaned_char = re.sub(r"[.\s]", "", char).lower()
                        if cleaned_char:
                            if k_idx < len(match) and cleaned_char == match[k_idx]:
                                k_idx += 1
                            else:
                                if k_idx == len(match):
                                    original_prefix_end_index = char_idx
                                    break
                    if k_idx == len(match) and original_prefix_end_index == 0:
                        original_prefix_end_index = len(name_clean)
    
                    rest = name_clean[original_prefix_end_index:].strip()
    
        ## 3. Split Name and Surname
        parts = rest.strip().split()
        if len(parts) >= 2:
            name = " ".join(parts[:-1])
            surname = parts[-1]
        elif len(parts) == 1:
            name, surname = parts[0], None
        else:
            name, surname = None, None
    
        return pd.Series([prefix_raw, name, surname])



    # Apply the extract_parts function
    result_df = df[name_column].apply(extract_parts)

    # Rename result columns to Thai
    result_df.columns = ['คำนำหน้า', 'ชื่อ', 'นามสกุล']

    # Combine into final DataFrame
    final_df = pd.concat([df, result_df], axis=1)

    return final_df

## 2nd function : to add issue text to another column
def add_issue(df, target_column, issuenote_column, match_values: set,note_txt):

    if issuenote_column not in df.columns:
     df[issuenote_column] = None
    
    def update_note(row):
        if row[target_column] in match_values:
            current = str(row[issuenote_column]).strip() if pd.notnull(row[issuenote_column]) else None
            return current + "/" + note_txt if current else note_txt
        return row[issuenote_column]
    
    df[issuenote_column] = df.apply(update_note, axis=1)
    return df