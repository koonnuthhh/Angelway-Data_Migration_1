import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
import os

# === Load reference data only once ===
def load_reference_data():
    base_path = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_path, "Thai_data.xlsx")


    province_content = pd.read_excel(excel_path, sheet_name="กำหนดจังหวัด", usecols="B,C")
    province_content.columns =['pid','v']
    PROVINCE_LIST =province_content.to_dict(orient="records")

    district_content = pd.read_excel(excel_path, sheet_name="กำหนดอำเภอ|เขต", usecols="B,C,D")
    district_content.columns =['pid','v','ppid']
    DISTRICT_LIST= district_content.to_dict(orient="records")
    
    subdistrict_content = pd.read_excel(excel_path, sheet_name="กำหนดตำบล|แขวง", usecols="B,C,E")
    subdistrict_content.columns =['pid','v','dpid']
    SUB_DISTRICT_LIST = subdistrict_content.to_dict(orient="records")


   
    return PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST


# === Normalize Amphoe ===
def normalize_district_name(name, province_name):
    name = name.strip()
    if name == "เมือง": 
        return f"เมือง{province_name}"
    return name


# === Province Extractor ===
def extract_province_from_amphoe(row, province_set):
    amphoe = str(row.get('อำเภอ', '')).strip()
    amphoe_cleaned = re.sub(r'^(อ\.|อำเภอ|อ)\s*', '', amphoe).strip()
    amphoe_cleaned = re.sub(r'^(จ\.|จ)\s*', '', amphoe_cleaned).strip()

    if amphoe_cleaned in province_set:
        return pd.Series({'อำเภอ': '-', 'จังหวัด': amphoe_cleaned})
    else:
        return pd.Series({'อำเภอ': amphoe_cleaned, 'จังหวัด': row.get('จังหวัด', '')})


# === Zipcode Checker ===
def check_zipcode(row, ref_data):
    PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST = ref_data
    province_name = str(row["จังหวัด"]).strip()
    district_name = str(row["อำเภอ"]).strip()
    sub_district_name = str(row["ตำบล"]).strip()


    province = next((p for p in PROVINCE_LIST if p["v"] == province_name), None)
    if not province:
        return "ไม่พบข้อมูลจังหวัด"
    ppid = province["pid"]

    normalized_district = normalize_district_name(district_name, province_name)
    district = next(
        (d for d in DISTRICT_LIST if d["ppid"] == ppid and (
            d["v"] == normalized_district or
            d["v"].startswith(district_name) or
            district_name in d["v"]
        )),
        None
    )
    if not district:
        return "ไม่พบข้อมูลอำเภอ"
    dpid = district["pid"]
    row["อำเภอ"] = normalized_district

    subdistrict = next(
        (s for s in SUB_DISTRICT_LIST if s["dpid"] == dpid and s["v"] == sub_district_name),
        None
    )
    if not subdistrict:
        return "ไม่พบข้อมูลตำบล"
    sub_pid = subdistrict["pid"]


    row["จังหวัด"] = ppid
    row["อำเภอ"] = dpid
    row["ตำบล"] = sub_pid
    return ""


# === Row Processor ===
def process_row(row, ref_data):
    result = check_zipcode(row, ref_data)
    row["ผลการตรวจสอบ"] = result
    return row



# === Main function ===
def main_location(filepath):
    PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST = load_reference_data()
    ref_data = (PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST)
    province_set = set([p["v"] for p in PROVINCE_LIST])
    df = pd.read_excel(filepath)
    #print(df_preview.columns)
    #df = pd.read_excel(filepath, usecols="K,R,S,T,U,V,W,X")
    df['รหัสลูกค้า'] = df['รหัสลูกค้า'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
    df['รหัส ปณ.'] = df['รหัส ปณ.'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
    df = df[~(df['รหัสลูกค้า'].isna() | (df['รหัสลูกค้า'].str.strip() == ''))]

    df['ที่อยู่'] = df['ที่อยู่'].str.replace(r'\s*ม\.\s*', ' หมู่ ', regex=True).str.strip()

    for label, pattern in {
        'ตำบล': r'(?:ตำบล\s*|ต\.)\s*\S.+',
        'ถนน': r'(?:ถนน\s*|ถ\.)\s*\S.+',
        'หมู่บ้าน': r'(?:บ้าน\s*|บ\.)\s*\S.+'
    }.items():
        mask = df['ที่อยู่'].str.contains(pattern, regex=True, na=False)
        df.loc[mask, label] = df.loc[mask, 'ที่อยู่'].str.extract(f'({pattern})', expand=False)
        df.loc[mask, 'ที่อยู่'] = df.loc[mask, 'ที่อยู่'].str.replace(pattern, '', regex=True).str.strip()

    df['อำเภอ'] = df['อำเภอ'].fillna('')
    amphoe_extracted = df['ตำบล'].str.extract(r'((?:อ\.|อำเภอ)\s*\S.+)', expand=False)
    df.loc[amphoe_extracted.notna(), 'อำเภอ'] = amphoe_extracted.dropna()
    df['ตำบล'] = df['ตำบล'].str.replace(r'(?:อ\.|อำเภอ)\s*\S.+', '', regex=True).str.replace(r'^(?:ต\.|ตำบล|ต)\s*', '', regex=True).str.strip()
    df['อำเภอ'] = df['อำเภอ'].str.replace(r'กิ่ง\s*(อ\.|อำเภอ|อ)\s*', 'อำเภอ', regex=True)
    
    
    # ✅ ใช้ apply แล้วเก็บผลลัพธ์ไว้ก่อน
    location_results = df.apply(lambda row: extract_province_from_amphoe(row, province_set), axis=1) 

    # ✅ ตรวจสอบว่ามีทั้ง 2 คอลัมน์
    assert set(['อำเภอ', 'จังหวัด']).issubset(location_results.columns), "Missing expected columns in output!"


    df[['อำเภอ', 'จังหวัด']] = location_results[['อำเภอ', 'จังหวัด']]

    if "ผลการตรวจสอบ" not in df.columns:
        df["ผลการตรวจสอบ"] = ""

    df = df.apply(lambda row: process_row(row, ref_data), axis=1)

    new_df = pd.DataFrame({
        "รหัสลูกค้า": df["รหัสลูกค้า"],
        "ประเภทที่อยู่": 1,
        "ใช้ส่งเอกสาร": "Y",
        "ที่อยู่": df["ที่อยู่"],
        "หมู่บ้าน/อาคาร": df.get("หมู่บ้าน", ""),
        "ซอย": np.nan,
        "ถนน": df.get("ถนน", ""),
        "ตำบล": df["ตำบล"],
        "อำเภอ": df["อำเภอ"],
        "จังหวัด": df["จังหวัด"],
        "รหัสไปรษณีย์": df["รหัส ปณ."],
        "โทรศัพท์มือถือ": df["โทรศัพท์ 1"],
        "สถานที่ใกล้เคียง": np.nan,
        "สถานะการอยู่อาศัย": np.nan,
        "ประเภทการอยู่อาศัย": np.nan,
        "ค่าเช่า/เดือน": np.nan,
        "ระยะเวลาการอยู่อาศัย/เดือน": np.nan,
        "ระยะเวลาการอยู่อาศัย/ปี": np.nan,
        "หมายเหตุ": df["ผลการตรวจสอบ"]
    })

    return new_df



def split_address_components(full_address):
    full_address = str(full_address).strip()

    # Extract postal code (5 digits at the end)
    zipcode_match = re.search(r'(\d{5})$', full_address)
    zipcode = zipcode_match.group(1) if zipcode_match else ''
    address_wo_zip = re.sub(r'\d{5}$', '', full_address).strip()

    # Extract province (last word before postcode)
    province_match = re.search(r'([ก-๙]+)\s*$', address_wo_zip)
    province = province_match.group(1) if province_match else ''
    address_wo_province = address_wo_zip.replace(province, '').strip()

    # Extract amphoe (อ. or อำเภอ)
    amphoe_match = re.search(r'(อ\.|อำเภอ)\s*([ก-๙]+)', address_wo_province)
    amphoe = amphoe_match.group(2) if amphoe_match else ''
    address_wo_amphoe = re.sub(r'(อ\.|อำเภอ)\s*[ก-๙]+', '', address_wo_province).strip()

    # Extract subdistrict (ต. or ตำบล)
    subdistrict_match = re.search(r'(ต\.|ตำบล)\s*([ก-๙]+)', address_wo_amphoe)
    subdistrict = subdistrict_match.group(2) if subdistrict_match else ''
    address_only = re.sub(r'(ต\.|ตำบล)\s*[ก-๙]+', '', address_wo_amphoe).strip()

    return pd.Series({
        'ที่อยู่ที่': address_only,
        'ตำบล': subdistrict,
        'อำเภอ': amphoe,
        'จังหวัด': province,
        'รหัส ปณ.': zipcode
    })

def process_address_file(filepath,column_address_name,worksheet="Sheet1"):
    print("Before load reference file")
    PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST = load_reference_data()
    ref_data = (PROVINCE_LIST, DISTRICT_LIST, SUB_DISTRICT_LIST)
    print("After load reference file")
    
    # Step 1: Load header row from openpyxl
    wb = load_workbook(filepath, data_only=True)
    ws = wb[worksheet]
    header_row_cells = ws[1]  # openpyxl is 1-based
    source_headers = [cell.value if cell.value is not None else f"Unnamed: {i}" 
                    for i, cell in enumerate(header_row_cells)]
    print("Headers from openpyxl:", source_headers)

    # Step 2: หา column name ที่มีคำว่า column_address_name
    address_column = next((col for col in source_headers if column_address_name in str(col)), None)

    if address_column is None:
        raise ValueError(f"Column not found: '{column_address_name}'")

    # Step 3: อ่านเฉพาะคอลัมน์นั้นโดยกำหนดชื่อคอลัมน์เอง
    df = pd.read_excel(filepath, header=None, skiprows=1, names=source_headers, usecols=[address_column])
    print(df.head())

    df = df.rename(columns={address_column: "ที่อยู่"})

    # Extract components from raw address string
    df[['ที่อยู่ที่', 'ตำบล', 'อำเภอ', 'จังหวัด', 'รหัส ปณ.']] = df['ที่อยู่'].apply(split_address_components)

    # Optional: clean up zip codes to ensure they’re 5 digits
    df['รหัส ปณ.'] = df['รหัส ปณ.'].apply(lambda x: str(x).zfill(5))

 

    df = df.apply(lambda row: process_row(row, ref_data), axis=1)

    # Now build the final format
    new_df = pd.DataFrame({
        "ที่อยู่": df["ที่อยู่ที่"],
        #"หมู่บ้าน/อาคาร": "",  # Optional: parse if needed
        "ซอย": "",
        "ถนน": "",
        "ตำบล": df["ตำบล"],
        "อำเภอ": df["อำเภอ"],
        "จังหวัด": df["จังหวัด"],
        "รหัสไปรษณีย์": df["รหัส ปณ."],
        "โทรศัพท์มือถือ": "",  # You can map this from another column if available
        "หมายเหตุ": df["ผลการตรวจสอบ"]
    })

    return new_df





def select_format(filepath):
    required_columns = ['ตำบล']

    # Try loading only the first few rows to check for column names
    df = pd.read_excel(filepath, nrows =0)
    if any(col in df.columns for col in required_columns):        
         # for first Excel format
         output_df = main_location(filepath)
    else:
       # for second Excel format
       output_df = process_address_file(filepath,'ที่อยู่')
    return output_df


def start(path):
 result_df = select_format(path)
 return result_df