import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook

def ChangeFormatMoney(source_path, workbook, column):
    # === File paths ===
    source_file = source_path

    # Save with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = source_path

    # === Load workbook and worksheet ===
    wb = load_workbook(source_file)
    ws = wb[workbook]

    # === Get headers ===
    header = [cell.value for cell in ws[1]]

    # === Get column index for 'รายได้ประจำ' ===
    try:
        income_col_index = header.index(column) + 1
    except ValueError:
        raise Exception("❌ Column "+column+" not found in header row.")

    # === Income cleaning function ===
    def extract_and_convert_to_monthly(value):
        if value is None or str(value).strip() == "":
            return "-"

        raw = str(value).strip().replace(" ", "")
        unit = 'unknown'

        match_paren = re.search(r'\((.*?)\)', raw)
        if match_paren:
            content = match_paren.group(1)
            if 'ปี' in content:
                unit = 'yearly'
            elif 'เดือน' in content:
                unit = 'monthly'

        if unit == 'unknown':
            if any(keyword in raw for keyword in ['/ปี', 'ต่อปี']):
                unit = 'yearly'
            elif any(keyword in raw for keyword in ['/เดือน', 'ต่อเดือน']):
                unit = 'monthly'

        if unit == 'unknown':
            if re.search(r'\d+(,?\d+)*[ ]*(ปี|ต่อปี)', raw):
                unit = 'yearly'
            elif re.search(r'\d+(,?\d+)*[ ]*(เดือน|ต่อเดือน)', raw):
                unit = 'monthly'

        match_num = re.search(r'\d[\d,\.]*', raw)
        if not match_num:
            return "-"

        raw_num = match_num.group(0).replace(",", "")
        try:
            num = float(raw_num)
        except:
            return "-"

        if num == 0:
            return "-"

        if unit == 'unknown':
            unit = 'yearly' if num >= 100000 else 'monthly'

        if unit == 'yearly':
            num = num / 12

        return f"{num:,.2f}"

    # === Process each row ===
    for row in ws.iter_rows(min_row=2):
        cell = row[income_col_index - 1]
        cleaned_monthly = extract_and_convert_to_monthly(cell.value)
        cell.value = cleaned_monthly

    # === Save ===
    wb.save(output_file)
    print("✅ Cleaned and saved to:\n", output_file)
