from openpyxl import load_workbook
from datetime import datetime
import os

def transform_guarantor_data(source_file: str, destination_file: str) -> str:
    """
    Copies and transforms data from a source Excel file into a WL template (1.3 or 2.3),
    mapping values appropriately and updating the destination file in-place.

    Additionally fills columns:
    - "ประเภทบัตร" with "1"
    - "ประเภทที่อยู่" with "1"
    - "ใช้ส่งเอกสาร" with "Y"
    for all rows that have any entries.

    Also combines 'ฟังก์ชันคู่ค้า' and 'เอกสาร' into 'หมายเหตุ' in format:
    "ฟังก์ชันคู่ค้า-เอกสาร"
    """

    template_column_mapping = {
        'Template1': {
            'รหัส': 'รหัสลูกค้า',
            'เลขที่บัตรประชาชน': 'เลขที่บัตร',
            'โทรศัพท์': 'โทรศัพท์มือถือ',
        },
        'Template2': {
            'รหัส': 'รหัสลูกค้า',
            'ที่อยู่': 'ที่อยู่ลูกค้า',
            'โทรศัพท์': 'Phone',
        }
    }

    source_wb = load_workbook(source_file, data_only=True)
    source_ws = source_wb.active
    source_header = [cell.value for cell in source_ws[1]]

    dest_wb = load_workbook(destination_file)
    template_sheet = None
    column_mapping = {}

    for sheet_name in ['Template1', 'Template2']:
        if sheet_name in dest_wb.sheetnames:
            template_sheet = dest_wb[sheet_name]
            column_mapping = template_column_mapping[sheet_name]
            break

    if template_sheet is None:
        raise Exception("❌ Template sheet not found (Template1 or Template2).")

    dest_header = [cell.value for cell in template_sheet[1]]

    def get_col_idx(col_name):
        try:
            return dest_header.index(col_name) + 1
        except ValueError:
            print(f"⚠️ Column '{col_name}' not found. Will skip filling it.")
            return None

    card_type_col_index = get_col_idx('ประเภทบัตร')
    address_type_col_index = get_col_idx('ประเภทที่อยู่')
    send_doc_col_index = get_col_idx('ใช้ส่งเอกสาร')
    remarks_col_index = get_col_idx('หมายเหตุ')

    # Indices for extra source columns
    try:
        function_col_idx = source_header.index("ฟังก์ชันคู่ค้า") + 1
        doc_col_idx = source_header.index("เอกสาร") + 1
    except ValueError as e:
        print(f"⚠️ Required source columns not found: {e}")
        function_col_idx = doc_col_idx = None

    # Map and copy values from source to destination
    for src_col, dest_col in column_mapping.items():
        if src_col not in source_header:
            print(f'⚠️ Source column "{src_col}" not found. Skipping.')
            continue

        try:
            dest_col_index = dest_header.index(dest_col) + 1
        except ValueError:
            print(f'⚠️ Destination column "{dest_col}" not found. Skipping.')
            continue

        src_col_index = source_header.index(src_col) + 1

        for row_idx in range(2, source_ws.max_row + 1):
            value = source_ws.cell(row=row_idx, column=src_col_index).value
            template_sheet.cell(row=row_idx, column=dest_col_index, value=value)

            if value is not None:
                if card_type_col_index:
                    template_sheet.cell(row=row_idx, column=card_type_col_index, value="1")
                if address_type_col_index:
                    template_sheet.cell(row=row_idx, column=address_type_col_index, value="1")
                if send_doc_col_index:
                    template_sheet.cell(row=row_idx, column=send_doc_col_index, value="Y")

        print(f'✅ Copied "{src_col}" → "{dest_col}"')

    # Add combined remark to "หมายเหตุ"
    if function_col_idx and doc_col_idx and remarks_col_index:
        for row_idx in range(2, source_ws.max_row + 1):
            function_val = source_ws.cell(row=row_idx, column=function_col_idx).value
            doc_val = source_ws.cell(row=row_idx, column=doc_col_idx).value
            if function_val and doc_val:
                combined_remark = f"{function_val}-{doc_val}"
                template_sheet.cell(row=row_idx, column=remarks_col_index, value=combined_remark)

        print(f'✅ Added combined remarks to "หมายเหตุ" column')

    dest_wb.save(destination_file)
    print(f'\n🎉 Done! File updated in place:\n{destination_file}')

    return destination_file
