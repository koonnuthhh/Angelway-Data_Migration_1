from openpyxl import load_workbook
from Levenshtein import distance as levenshtein_distance
import re

def clean_name(exel_file, name_column="ชื่อ-สกุล", sheet_name=None):
    wb = load_workbook(exel_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find header row and target column index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if name_column not in headers:
        raise ValueError(f"Column '{name_column}' not found in Excel headers: {headers}")
    name_col_idx = headers.index(name_column) + 1  # openpyxl is 1-based

    thai_english_prefixes = [
        "เด็กชาย", "ด.ช.","เด็กหญิง", "ด.ญ.","นาวสาว",
        "นาย", "นาง", "นางสาว", "น.ส.", "าง", "คุณ", "นาน",
        "ดอกเตอร์", "ดร.", "ศาสตราจารย์", "ศ.", "ผู้ช่วยศาสตราจารย์", "ผศ.", "รองศาสตราจารย์", "รศ.", "ศาสตราจารย์ ดร.", "ศ.ดร.", "ผศ.ดร.", "รศ.ดร.",
        "พลเอก", "พล.อ.", "พลโท", "พล.ท.", "พลตรี", "พล.ต.", "พันเอก", "พ.อ.", "พันโท", "พ.ท.", "พันตรี", "พ.ต.", "ร้อยเอก", "ร.อ.", "ร้อยโท", "ร.ท.",
        "ร้อยตรี", "ร.ต.", "ร.ต.ต.", "ร.ต.อ.", "สิบตรี", "ส.ต.", "พลตำรวจเอก", "พล.ต.อ.", "พันตำรวจเอก", "พ.ต.อ.", "จ.ส.อ.", "ด.ต.", "ส.ท.", "จ.ส.ต.",
        "ว่าที่ ร.ต.หญิง", "ส.ณ.", "ส.อ.", "จ.อ.", "ดาบตำรวจ",
        "หม่อม",  "หม่อมหลวง", "ม.ล.", "หม่อมราชวงศ์", "ม.ร.ว.", "หม่อมเจ้า", "ม.จ.", "คุณหญิง", "คุณชาย", "ท่านผู้หญิง",
        "นายแพทย์", "นพ.", "ทันตแพทย์", "ทพ.", "เภสัชกร", "ภก.", "ภกญ.", "สัตวแพทย์", "สพ.", "วิศวกร", "วศ.", "นิติกร", "นก.", "บริษัท", "หจก.", "ห้างหุ้นส่วนจำกัด",
        "พระ",
        "Mr.", "Mrs.", "Ms.", "Miss", "Dr.", "Prof.",  "Sir", "Madam", "Lady", "Lord", "Rev.", "Capt.",  "Col.", "Gen.", "Lt.", "Sgt.", "Major"
    ]

    all_prefixes = []
    for p in thai_english_prefixes:
        all_prefixes.append(p)
        if not p.startswith("ว่าที่"):
            all_prefixes.extend([
                "ว่าที่" + p,
                "ว่าที่ " + p,
                "ว่าที" + p,
                "ว่าที " + p
            ])

    prefix_map = {
        p.replace(".", "").replace(" ", "").lower(): p for p in all_prefixes
    }
    prefix_keys_sorted = sorted(prefix_map.keys(), key=lambda x: -len(x))

    def extract_parts(full_name):
        if full_name is None or not isinstance(full_name, str):
            return None, None, None

        name_clean = full_name.strip()
        name_for_match = re.sub(r"[.\s]", "", name_clean).lower()

        prefix_raw = None
        rest = name_clean

        for key in prefix_keys_sorted:
            if name_for_match.startswith(key):
                prefix_raw = prefix_map[key]
                rest = name_clean[len(prefix_raw):].strip()
                break

        if prefix_raw is None:
            best_match = ("", float("inf"))
            for i in range(2, min(20, len(name_for_match) + 1)):
                snippet = name_for_match[:i]
                for prefix_key in prefix_keys_sorted:
                    dist = levenshtein_distance(snippet, prefix_key)
                    if dist < best_match[1]:
                        best_match = (prefix_key, dist)

            match, dist = best_match
            if dist <= 2 and name_for_match.startswith(match):
                prefix_raw = prefix_map[match]
                rest = name_clean[len(prefix_raw):].strip()

        parts = rest.split()
        if len(parts) >= 2:
            name = " ".join(parts[:-1])
            surname = parts[-1]
        elif len(parts) == 1:
            name, surname = parts[0], None
        else:
            name, surname = None, None

        return prefix_raw, name, surname

    # Process each row
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = row[name_col_idx - 1]
        prefix, name, surname = extract_parts(full_name)
        results.append({
            name_column: full_name,
            'คำนำหน้า': prefix,
            'ชื่อ': name,
            'นามสกุล': surname
        })

    return results  # You can convert this to a DataFrame if needed
