import pandas as pd
import xlrd
from openpyxl import load_workbook

def dowload_df(sourcepath, sheet_index=0, header_row=0):
    from openpyxl import load_workbook
    import pandas as pd

    try:
        # Use openpyxl to read the sheet
        wb = load_workbook(sourcepath, data_only=True)
        ws = wb.worksheets[sheet_index]
        data = list(ws.values)
        header = data[header_row]
        rows = data[header_row + 1:]
        df = pd.DataFrame(rows, columns=header)
        print(f"✅ Download {sourcepath} as XLSX (openpyxl) success!!")
        # --- Universal header fallback ---
        if all(str(col).startswith('Unnamed') or col is None for col in df.columns):
            print("⚠️ All columns are Unnamed. Trying to auto-detect header row...")
            # Scan first 10 rows for a row with all non-None values
            for i in range(min(10, len(data))):
                row = data[i]
                if row and all(cell is not None and str(cell).strip() != '' for cell in row):
                    print(f"✅ Found likely header row at index {i}: {row}")
                    header = [str(cell).strip() for cell in row]
                    rows = data[i + 1:]
                    df = pd.DataFrame(rows, columns=header)
                    break
        return df
    except Exception as e:
        print(f"⚠️ openpyxl read failed: {e}\nFallback to pandas/xlrd/csv...")
        # Fallback to your original logic
        try:
            dowload_file = pd.read_excel(sourcepath, sheet_name=sheet_index)
            # --- Universal header fallback for pandas ---
            if all(str(col).startswith('Unnamed') or col is None for col in dowload_file.columns):
                print("⚠️ All columns are Unnamed (pandas). Trying to auto-detect header row...")
                # Try reading first 10 rows to find header
                preview = pd.read_excel(sourcepath, sheet_name=sheet_index, header=None, nrows=10)
                for i, row in preview.iterrows():
                    if all(cell is not None and str(cell).strip() != '' for cell in row):
                        print(f"✅ Found likely header row at index {i}: {list(row)}")
                        df = pd.read_excel(sourcepath, sheet_name=sheet_index, header=i)
                        return df
            return dowload_file
        except Exception as e:
            print(f"The format is not xlsx.\n Changing read method to xls...")
            try:
                import xlrd
                workbook = xlrd.open_workbook(sourcepath)
                sheet = workbook.sheet_by_index(sheet_index)
                headers = sheet.row_values(0)
                data = [sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows)]
                print("Dowload {sourcepath} success!!")
                return pd.DataFrame(data, columns=headers)
            except Exception as e:
                print(f"The format is not xls as well.\n Changing read method to csv(txt)...")
                df = pd.read_csv(sourcepath, delimiter='\t',  encoding='cp874')
                print("Dowload {sourcepath} success!!")
                return df
     
def dowload_df_filename(sourcepath, sheet_name, source_header_row=0):
    try:
        # --- Use openpyxl to get actual headers from source_header_row ---
        wb = load_workbook(sourcepath, data_only=True)
        ws = wb[sheet_name]
        header_row_cells = ws[source_header_row + 1]  # openpyxl is 1-based
        source_headers = [
            cell.value if cell.value is not None else f"Unnamed: {i}"
            for i, cell in enumerate(header_row_cells)
        ]

        # --- Read data below header row ---
        df = pd.read_excel(
            sourcepath,
            sheet_name=sheet_name,
            header=None,
            skiprows=source_header_row + 1,
            engine='openpyxl'
        )

        # --- Adjust column headers to match data width ---
        num_data_columns = df.shape[1]
        num_header_columns = len(source_headers)

        if num_header_columns != num_data_columns:
            print(f"⚠️ Header/Data column mismatch: {num_header_columns} headers vs {num_data_columns} columns")

            if num_header_columns > num_data_columns:
                # Truncate headers
                source_headers = source_headers[:num_data_columns]
            else:
                # Add unnamed headers
                source_headers += [f"Unnamed: {i}" for i in range(num_header_columns, num_data_columns)]

        df.columns = [str(h).strip() for h in source_headers]
        print(f"✅ Download {sourcepath} as XLSX (openpyxl) success!!")
        return df

    except Exception as e:
        print(f"⚠️ XLSX read failed: {e}\nFallback to .xls...")

        try:
            workbook = xlrd.open_workbook(sourcepath)
            sheet = workbook.sheet_by_name(sheet_name)
            headers = sheet.row_values(source_header_row)
            data = [sheet.row_values(row_idx) for row_idx in range(source_header_row + 1, sheet.nrows)]

            # Ensure headers match data
            if headers and data and len(headers) != len(data[0]):
                print(f"⚠️ Header/Data mismatch in .xls: fixing...")
                if len(headers) < len(data[0]):
                    headers += [f"Unnamed: {i}" for i in range(len(headers), len(data[0]))]
                else:
                    headers = headers[:len(data[0])]

            print(f"✅ Download {sourcepath} as XLS success!!")
            return pd.DataFrame(data, columns=headers)

        except Exception as e:
            print(f"⚠️ XLS read failed: {e}\nFallback to CSV/TXT...")

            try:
                df = pd.read_csv(sourcepath, delimiter='\t', encoding='cp874')
                print(f"✅ Download {sourcepath} as CSV/TXT success!!")
                return df

            except Exception as e:
                print(f"❌ Failed to read file in all formats: {e}")
                return None
