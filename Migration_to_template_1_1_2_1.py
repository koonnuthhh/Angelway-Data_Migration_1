import pandas as pd
from function.FunctionForX_1AndX_2AndX_3.MappingCleanData import update_xlsx_with_clean_data
from function.FunctionForX_1AndX_2AndX_3.clean_name import clean_name
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from function.FunctionForX_1AndX_2AndX_3.ChangeFormatMoney import ChangeFormatMoney
from function.FunctionForX_1AndX_2AndX_3.define_customer_type import define_customer_type
from function.FunctionForX_1AndX_2AndX_3.setWL_location import start
from datetime import datetime
import os

def process_wl_data(
    source_file,
    source_sheet,
    template_1_path,
    template_1_sheet,
    template_2_path,
    template_2_sheet
):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_1 = template_1_path
    output_2 = template_2_path

    df_clean = clean_name(source_file, "ชื่อลูกค้า")
    column_map = {'คำนำหน้า': 'คำนำหน้า', 'ชื่อ': 'ชื่อ', 'นามสกุล': 'นามสกุล'}

    update_xlsx_with_clean_data(source_file, df_clean, column_map, sheet_name=source_sheet)

    def clean(text):
        if pd.isna(text): return ''
        return str(text).strip().replace('\u200b', '').replace('\xa0', ' ').replace('\t', '').replace('\n', '').lower()

    def format_date(date):
        try:
            if pd.isna(date) or str(date).strip() == '':
                return ''
            parsed_date = pd.to_datetime(date, errors='coerce')
            if parsed_date is not pd.NaT:
                return parsed_date.strftime('%d/%m/%Y')
            return ''
        except Exception:
            return ''

    source_df_raw = pd.read_excel(source_file, sheet_name=source_sheet, dtype=str)
    source_df = source_df_raw.copy()
    source_df.columns = [clean(col) for col in source_df.columns]

    def process_template_1():
        column_mapping_1 = {
            'คำนำหน้า': 'คำนำหน้า',
            'ชื่อ': 'ชื่อ',
            'นามสกุล': 'นามสุกล',
            'รหัสลูกค้า': 'รหัสลูกค้า',
            'เพศ': 'เพศ',
            'เลขที่บัตรปชช./ภาษี': 'เลขที่บัตร',
            'รายได้ประจำ': 'รายได้ประจำ',
            'สถานะภาพสมรส': 'สถานะภาพสมรส',
            'วันเกิด': 'วันเกิด',
            'อาชีพ': 'อาชีพ/ประเภทธุรกิจ',
            'โทรศัพท์ 1': 'โทรศัพท์มือถือ'
        }

        wb = load_workbook(template_1_path)
        ws = wb.active
        header = [clean(cell.value) for cell in ws[1]]
        column_mapping_clean = {clean(k): v for k, v in column_mapping_1.items()}

        for src_col_clean, dest_col in column_mapping_clean.items():
            if dest_col not in header or src_col_clean not in source_df.columns:
                continue
            col_index = header.index(dest_col) + 1

            if dest_col == 'โทรศัพท์มือถือ':
                col1, col2 = clean('โทรศัพท์ 1'), clean('โทรศัพท์ 2')
                values = source_df.apply(lambda row: row.get(col1) or row.get(col2) or '', axis=1)
            elif dest_col == 'วันเกิด':
                values = source_df[src_col_clean].apply(format_date)
            else:
                values = source_df[src_col_clean].apply(clean)

            for row_idx, value in enumerate(values, start=2):
                ws.cell(row=row_idx, column=col_index, value=value).alignment = Alignment(horizontal='center', vertical='center')

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=11, value='1').alignment = Alignment(horizontal='center', vertical='center')

        columns_to_fill = ['ชื่อเล่น', 'รหัสเกรดลูกค้า', 'รหัสกลุ่มลูกค้า', 'อายุ(ปี)', 'วันที่ออกบัตร', 'วันที่บัตรหมดอายุ',
                           'เชื้อชาติ', 'สัญชาติ', 'ระดับการศึกษา', 'ตำแหน่ง', 'ชื่อสถานที่ทำงาน',
                           'อายุงาน(ปี)', 'อายุงาน(เดือน)', 'รายได้อื่นๆ', 'แหล่งที่มารายได้',
                           'ค่าใช้จ่าย/เดือน', 'โทรศัพท์บ้าน', 'อีเมล์', 'บุคคลอ้างอิง', 'ความสัมพันธ์',
                           'เบอร์โทรบุคคลอ้างอิง', 'สาขาภาษี', 'หมายเหตุ']
        for col_name in columns_to_fill:
            col_name_clean = clean(col_name)
            if col_name_clean in header:
                col_index = header.index(col_name_clean) + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_index).alignment = Alignment(horizontal='center', vertical='center')

        wb.save(output_1)
        print('✅ Template 1 บันทึกที่: ' + output_1)

    def process_template_2():
        column_mapping_2 = {
            'คำนำหน้า': 'คำนำหน้า',
            'ชื่อ': 'ชื่อ',
            'นามสกุล': 'สกุล',
            'โทรศัพท์ 1': 'โทรศัพท์มือถือ',
            'ที่อยู่': 'ที่อยู่',
            'หมู่บ้าน': 'หมู่บ้าน/อาคาร',
            'รหัสลูกค้า': 'รหัสลูกค้า',
            'รหัส ปณ.': 'รหัสไปรษณีย์'
        }

        wb = load_workbook(template_2_path)
        ws = wb.active
        header = [clean(cell.value) for cell in ws[1]]
        column_mapping_clean = {clean(k): v for k, v in column_mapping_2.items()}

        for src_col_clean, dest_col in column_mapping_clean.items():
            if dest_col not in header or src_col_clean not in source_df.columns:
                continue
            col_index = header.index(dest_col) + 1

            if dest_col == 'โทรศัพท์มือถือ':
                col1, col2 = clean('โทรศัพท์ 1'), clean('โทรศัพท์ 2')
                values = source_df.apply(lambda row: row.get(col1) or row.get(col2) or '', axis=1)
            else:
                values = source_df[src_col_clean].fillna('')

            for row_idx, value in enumerate(values, start=2):
                ws.cell(row=row_idx, column=col_index, value=value).alignment = Alignment(horizontal='center', vertical='center')

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=2, value=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=3, value='Y').alignment = Alignment(horizontal='center', vertical='center')

        columns_to_fill = ['ซอย', 'ถนน', 'สถานที่ใกล้เคียง',
                           'สถานะการอยู่อาศัย', 'ประเภทการอยู่อาศัย',
                           'ค่าเช่า/เดือน', 'ระยะเวลาการอยู่อาศัย/เดือน', 'ระยะเวลาการอยู่อาศัย/ปี']
        for col_name in columns_to_fill:
            col_name_clean = clean(col_name)
            if col_name_clean in header:
                col_index = header.index(col_name_clean) + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_index).alignment = Alignment(horizontal='center', vertical='center')

        wb.save(output_2)
        print('✅ Template 2 บันทึกที่: ' + output_2)

    process_template_1()
    process_template_2()

    ChangeFormatMoney(template_1_path, template_1_sheet, "รายได้ประจำ")
    define_customer_type(template_1_path, template_1_sheet, "คำนำหน้า", "ประเภทลูกหนี้")

    maw = start(source_file)
    print(maw)

    column_location_map = {
        'ซอย': 'ซอย',
        'ถนน': 'ถนน',
        'ตำบล': 'ตำบล',
        'อำเภอ': 'อำเภอ',
        'จังหวัด': 'จังหวัด',
        'รหัสไปรษณีย์': 'รหัสไปรษณีย์'
    }

    update_xlsx_with_clean_data(
        template_2_path,
        maw,
        column_location_map,
        sheet_name=template_2_sheet
    )

#Example Usage
#process_wl_data(
#    source_file=r"...",
#    source_sheet="Sheet1",
#    template_1_path=r"...",
#    template_1_sheet="...",
#    template_2_path=r"...",
#    template_2_sheet="..."
#)