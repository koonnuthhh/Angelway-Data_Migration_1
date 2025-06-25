import pandas as pd
import os
import sys
import function.change_Branchcode7_9.branchcode_change as bc
import xlrd
from Migration_to_Template_9_WL import dowload_df
from openpyxl import load_workbook
from function.ColumnMappingFunction import map_excel_columns 
import function.Function_count_group_Wl7.transferGroup as tg
output_file = "Template_7_WL_output.xlsx"

def Migration_to_Template_7_WL(source_file_1,source_sheet_1,source_file_2,source_sheet_2,destination_file, Template_11_file, Template_9_file, aging_file,destination_sheet) :

    # First mapping (Source 1 into Destination)
    mapped_df1 = map_excel_columns(
        source_file_1,
        destination_file,
        source_sheet_1,
        destination_sheet,
        column_mapping={   # Source 1 mapping
            'cont_no': 'cont_no',
            'cont_date': 'cont_date',
            'cust_code': 'cust_code',
            'Cust_ID': 'Cust_ID',
            'collateral_type': 'collateral_type',
            'chassis_no': 'chassis_no',
            'send_bill_status': 'send_bill_status',
            'collateral_vat_rate': 'collateral_vat_rate',
            'product_price': 'product_price',
            'net_product_price': 'net_product_price',
            'vat_product_price': 'vat_product_price',
            'down_payment': 'down_payment',
            'principal': 'principal',
            'net_principal': 'net_principal (H)',
            'vat_principal': 'vat_principal',
            'interest_year': 'interest_flat_rate',
            'period': 'period',
            'installment': 'installment',
            'net_installment': 'net_installment',
            'vat_installment': 'vat_installment',
            'installment_amount': 'installment_amount (G)',
            'net_installment_amount': 'net_installment_amount',
            'vat_installment_amount': 'vat_installment_amount (J)',        
            'penalty_method': 'penalty_method',
            'penalty_rate': 'penalty_rate',
            'material_group': 'material_group',
            'deferred_interest': 'deferred_interest (I)',
        }
    )

    # Second mapping (Source 2 into Destination)
    mapped_df2 = map_excel_columns(
        source_file_2,
        destination_file,  # Same template to get headers
        source_sheet_2,
        destination_sheet,
        column_mapping={   # Source 2 mapping
            'PRC_GW': 'standard_price',
            'NET_PRC_GW': 'net_standard_price',
            'VAT_PRC_GW': 'vat_standard_price',
            'period_installment': 'period_installment',
            'first_due_date': 'first_due_date',
            'last_due_date': 'last_due_date',
            'sale_code': 'sale_code',
            'billcollector_code': 'billcollector_code',
        }
    )

    # ⚠️ Now MERGE the two mapped DataFrames:
    # Where mapped_df2 is not empty, fill mapped_df1 (destination)
    final_df = mapped_df1.copy()

    for column in mapped_df2.columns:
        if column in final_df.columns:
            # Only replace if mapped_df2 has non-NaN (not empty) values
            final_df[column] = mapped_df2[column].combine_first(final_df[column])

    # Fill the column in the destination DataFrame
    if "checker_code" in final_df.columns:
        final_df["checker_code"] = "4261"
    else:
        final_df["checker_code"] = "4261"


    # Fill the column in the destination DataFrame
    if "penalty_late" in final_df.columns:
        final_df["penalty_late"] = "7"
    else:
        final_df["penalty_late"] = "7"



    # ✅ Export the final merged dataframe
    df = pd.read_excel(source_file_1,usecols='G,I',dtype=str)
    df.columns =['cont_group','branch_code']
    ######logic import file final_df= logic
    df = tg.start(df)
    df = bc.start(df)
    final_df['cont_group'] = df['cont_group']
    final_df['branch_code'] = df['branch_code']

    print(df.columns)
    print(final_df.columns)
    final_df['net_installment_amount_issue']=final_df['net_installment']*final_df['period']
    final_df['vat_installment_amount_issue']=final_df['vat_installment']*final_df['period']
    final_df['เช็ค Diff กับช่อง AC']=final_df['deferred_interest (I)'] + final_df['vat_installment_amount (J)']+final_df['net_principal (H)']

    final_df['deferred_interest_issue']=final_df['net_installment_amount_issue'] - final_df['net_principal (H)']
    final_df['เช็ค AC-AQ'] = final_df['installment_amount (G)'] - final_df['เช็ค Diff กับช่อง AC']
    final_df['เช็ค Diff กับช่อง AC.1'] = final_df['net_principal (H)']+final_df['vat_installment_amount_issue']+final_df['deferred_interest_issue']
    final_df['เช็ค AC-AS'] = final_df['installment_amount (G)']-final_df['เช็ค Diff กับช่อง AC.1']
    # add  more requirement
    final_df['ภาษีเช่าซื้อทั้งหมด'] = final_df['net_installment_amount']*(7/100)
    final_df['ตรวจ_diff_ภาษีซื้อ'] = final_df['vat_installment_amount (J)'] -  final_df['ภาษีเช่าซื้อทั้งหมด']
    
    # dowload template 11 and aging file
    wb = load_workbook(Template_11_file, data_only=True)
    # ระบุชื่อ Sheet (หรือใช้ sheet แรก)
    sheet = wb.active  # หรือ wb['ชื่อชีต']
    # ดึงข้อมูลใน Sheet
    data = list(sheet.values)
    # แยก header กับ data
    header = data[0]
    rows = data[1:]
    # แปลงเป็น DataFrame
    Template_11_df = pd.DataFrame(rows, columns=header)
    
    
    
    wb = load_workbook(Template_9_file, data_only=True)

    # ระบุชื่อ Sheet (หรือใช้ sheet แรก)
    sheet = wb.active  # หรือ wb['ชื่อชีต']

    # ดึงข้อมูลใน Sheet
    data = list(sheet.values)

    # แยก header กับ data
    header = data[0]
    rows = data[1:]

    # แปลงเป็น DataFrame
    Template_9_df = pd.DataFrame(rows, columns=header)
    
    #ลบ column ที่มีชื่อซ้ำ
    Template_9_df = Template_9_df.loc[:, ~Template_9_df.columns.duplicated()]
    print("Dowload template 9 success!!")
    
    

    Aging_df = dowload_df(aging_file)
    
    
    # Merge the two DataFrames on 'cont_no'
    vat_payment_sum = Template_11_df.groupby('cont_no')['vat_payment'].sum()
    
    unique_template_11 = Template_11_df.drop_duplicates(subset=['cont_no']).copy()
    
    unique_template_11['vat_payment'] = unique_template_11['cont_no'].map(vat_payment_sum)
    
    merged_df = pd.merge(final_df, unique_template_11[['cont_no', 'vat_payment']], on='cont_no', how='inner')
    
    

    # Apply the logic only to the merged DataFrame
    merged_df['ภาษีรอตัด'] = merged_df['vat_installment_amount (J)'] - merged_df['vat_payment']
    
    filtered = Template_9_df[Template_9_df['payfor_code'] != 1001]
    
    # คำนวณผลรวม payment ของแต่ละ cont_no ใน filtered (จะได้ Series ที่ index เป็น cont_no)
    payment_sum = filtered.groupby('cont_no')['payment'].sum()

    # จากนั้นเอาไปแมปใน Template_9_df หรือใช้ merge
    unique_template_9 = Template_9_df.drop_duplicates(subset=['cont_no']).copy()

    # แมปผลรวมลงใน unique_template_9
    unique_template_9['payment_sum'] = unique_template_9['cont_no'].map(payment_sum)

    #sprint(unique_template_9[['cont_no', 'payment_sum']])
    
    
    
    
    remaining_debt = final_df.merge(unique_template_9[['cont_no', 'payment_sum']], on='cont_no')
    remaining_debt['มูลหนี้คงเหลือ'] = remaining_debt['installment_amount (G)'] - remaining_debt['payment_sum']
    
    # remaining_debt.to_excel(   
    #     r'remaining_debt.xlsx',
    #     index=False,
    # )

    final_df = pd.merge(final_df, merged_df[['cont_no', 'ภาษีรอตัด']], on='cont_no', how='left')
    final_df = pd.merge(final_df, remaining_debt[['cont_no', 'มูลหนี้คงเหลือ']], on='cont_no', how='left')
    final_df = pd.merge(final_df, Aging_df[['เลขที่สัญญา', ' ผลรวมรายการเปิ']], left_on='cont_no', right_on='เลขที่สัญญา', how='left')
    final_df.drop(columns=['เลขที่สัญญา'], inplace=True)
    final_df.rename(columns={' ผลรวมรายการเปิ': 'มูลหนี้คงเหลือตาม_Aging'}, inplace=True)
    final_df['เช็ค_diff_มูลหนี้คงเหลือ'] = final_df['มูลหนี้คงเหลือ'] - final_df['มูลหนี้คงเหลือตาม_Aging']
    
    
    
    
    print("saving file...")
    final_df.to_excel(
        output_file,
        index=False,
        sheet_name="ข้อมูลสัญญาเชื้อ"
    )


# source_file_1 = "D:/Angelway/Migration to python/File_testing/WL_test/Tem.7/wl_zbfamt_04.2025(Temp7.1).XLSX"
# source_sheet_1 = "Sheet1"
# source_file_2 = "D:/Angelway/Migration to python/File_testing/WL_test/Tem.7/wl_zbfamt2_04.2025 (Temp7.2).XLSX"
# source_sheet_2 = "Sheet1"
# destination_file = "D:/Angelway/Migration to python/File_testing/WL_test/Tem.7/7-ข้อมูลสัญญาเช่าซื้อ.xlsx"
# Template_11_file = "D:/Angelway/Migration to python/MigrationFunction/Angelway-Data_Migration_1/Template_11_WL_output.xlsx"
# Template_9_file = "D:/Angelway/Migration to python/MigrationFunction/Angelway-Data_Migration_1/Template_9_WL_output.xlsx"
# aging_file = "D:/Angelway/Migration to python/File_testing/WL_test/Tem.7/Aging_04.2025.xls"
# destination_sheet = "ข้อมูลสัญญาเชื้อ"
# Migration_to_Template_7_WL(source_file_1,source_sheet_1,source_file_2,source_sheet_2,destination_file, Template_11_file, Template_9_file, aging_file,destination_sheet)