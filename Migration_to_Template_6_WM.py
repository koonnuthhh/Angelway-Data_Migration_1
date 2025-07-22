from openpyxl import load_workbook
import numpy as np
import os
import sys
from function.WM_clean_zfloan20 import WM_clean_zfloan20
import xlrd
from Migration_to_Template_9_WL import dowload_df
import math

output_file = "6-ข้อมูลสัญญาเงินกู้ลดต้นลดดอก_result.xlsx"

def Migration_to_Template_6_WM(source_file1,zfloan_raw,source_file3,source_file4,Temp9_File,Aging_file,destination_file) :
    import pandas as pd
    
    #Clean zfloan20 first
    WM_clean_zfloan20(zfloan_raw)
    
    # Declare variable
    source_file2 = "WM_Temp6_cleaned.xlsx" # zfloan20 กรอง 1 
    source_file2_2 = "WM_Temp6_cleanedV2.xlsx" # zfloan20 กรอง 2
    
    # หน้า sheet ของไฟล์ต้นฉบับ
    source_sheet1 = "รายละเอียด" # WM_1014100
    target_sheet = "ข้อมูลสัญญาลดต้นลดดอก" 
    
    # ชื่อ column เลขที่สัญญา ของไฟล์ engine
    id_column = "เลขที่สัญญา"

    ## ตัวแปรสำหรับกรอง zf20 ดิบ
    # กรองรอบ 1
    doc_col = "เลขเอกสาร" # reference เอกสาร
    clrng_col = "Clrng doc." # ใบแจ้งหนี้
    date_col = "การหักล้าง" # วันที่
    # กรองรอบ 2
    date_col_2 = "Pstng Date" # column วันที่สำหรับกรองรอบ 2
    date = "30/04/2025" # วันที่ที่จะกรองออก

    ## ตัวแปรสำหรับ zf60

    # --- Auto-select engine for Excel file ---
    file_ext = os.path.splitext(source_file1)[1].lower()
    if file_ext == ".xls":
        engine = "xlrd"
    else:
        engine = "openpyxl"

    try:
        # --- Read and clean WM ---
        WM = pd.read_excel(source_file1, sheet_name=source_sheet1, engine=engine, header=0)
        WM.columns = WM.columns.str.strip()
        print("WM columns:", list(WM.columns))
    except ValueError as e:
        # Sheet not found, print available sheet names
        print(f"Sheet '{source_sheet1}' not found in {source_file1}.")
        import xlrd
        if engine == "xlrd":
            book = xlrd.open_workbook(source_file1)
            print("Available sheets:", book.sheet_names())
        else:
            import openpyxl
            wb = openpyxl.load_workbook(source_file1, read_only=True)
            print("Available sheets:", wb.sheetnames)
        raise

    WM = WM[["เลขที่สัญญา","สาขา","ผลรวมรายการเปิด"]].copy()
    WM = WM.rename(columns={
        'สาขา': 'branch_code',
        'ผลรวมรายการเปิด': 'current_balance'
    })
    
    ## load data
    # --- Read and clean zf20_1 ---
    zf20_1 = pd.read_excel(source_file2, engine="openpyxl")
    zf20_1.columns = zf20_1.columns.str.strip()
    print("zf20_1 columns:", list(zf20_1.columns))
    # Use the correct, stripped column names for zf20_1
    zf20_1 = zf20_1.rename(columns={'เลขที่สัญญ': 'เลขที่สัญญา'})
    zf20_1 = zf20_1[["เลขที่สัญญา","ดบ.ในงวด","วันครบกำหน"]].copy()
    zf20_1 = zf20_1.rename(columns={
        'ดบ.ในงวด': 'interest_accrued',
        'วันครบกำหน': 'next_due_date'
    })

    ## load data
    # --- Read and clean zf20_2 ---
    zf20_2 = pd.read_excel(source_file2_2, engine="openpyxl")
    zf20_2.columns = zf20_2.columns.str.strip()
    print("zf20_2 columns:", list(zf20_2.columns))
    # Use the correct, stripped column names for zf20_2
    zf20_2 = zf20_2.rename(columns={'เลขที่สัญญ': 'เลขที่สัญญา'})
    zf20_2 = zf20_2[["เลขที่สัญญา","เงินต้น","ดบ.ในงวด",
                                "มูลค่าในงวด","@วันในงวด"]].copy()
    zf20_2 = zf20_2.rename(columns={
        'เงินต้น': 'installment_current',
        'ดบ.ในงวด': 'installment_interest',
        'มูลค่าในงวด': 'installment_over_due',
        '@วันในงวด': 'dpd'
    })

    ## load data
    # --- Read and clean zf60 ---
    zf60 = pd.read_excel(source_file4, engine="openpyxl")
    zf60.columns = zf60.columns.str.strip()
    print("zf60 columns:", list(zf60.columns))
    zf60['สถานะสัญญา'].astype

    # Example: Clean column 'col' to keep only leading digits as string
    zf60['สถานะสัญญา'] = zf60['สถานะสัญญา'].str.extract(r'ค้างชำระ\s+(\d+)')[0]
    zf60['สถานะสัญญา'] = pd.to_numeric(zf60['สถานะสัญญา'], errors='coerce').astype('Int64')
    zf60['สถานะสัญญา'] = zf60['สถานะสัญญา'].astype(str).replace('<NA>', '')

    zf60 = zf60[["สัญญา","สถานะสัญญา"]].copy()
    zf60.rename(columns={'สัญญา': 'เลขที่สัญญา'}, inplace=True)
    zf60 = zf60.rename(columns={'สถานะสัญญา': 'over_due_period'})

    ## zf50
    # --- Read and clean zf50 ---
    zf50 = pd.read_excel(source_file3, engine="openpyxl")
    zf50.columns = zf50.columns.str.strip()
    print("zf50 columns:", list(zf50.columns))

    # If all columns are Unnamed, try reading with openpyxl directly
    if all(col.startswith('Unnamed') for col in zf50.columns):
        import openpyxl
        wb = openpyxl.load_workbook(source_file3)
        ws = wb.active
        data = list(ws.values)
        # Find the first row with all non-None values (likely the header)
        for i, row in enumerate(data):
            if row and all(cell is not None for cell in row):
                header_row = i
                break
        else:
            header_row = 0
        headers = [str(cell).strip() for cell in data[header_row]]
        df_data = data[header_row+1:]
        zf50 = pd.DataFrame(df_data, columns=headers)
        zf50.columns = zf50.columns.str.strip()
        print("zf50 columns (openpyxl):", list(zf50.columns))

    zf50['วันเริ่มต้น'] = pd.to_datetime(zf50['วันเริ่มต้น'], errors='coerce')
    zf50['payment_day'] = zf50['วันเริ่มต้น'].dt.day
    
    zf50 = zf50[["เลขที่สัญญา","วันทำสัญญา","ลูกค้า","เลขที่จดทะเบียน VAT","หมายเลขถัง",
         "รหัสประเภทสินเชื่อ","ประเภทสินเชื่อ","เงินต้น","อัตราดอกเบี้ย", "จำนวนงวด",
         "ค่างวด","วันเริ่มต้น","วันสิ้นสุด","payment_day"]].copy()

    for col in ["เลขที่สัญญา", "ลูกค้า", "เลขที่จดทะเบียน VAT"]:
        zf50[col] = zf50[col].apply(lambda x: str(int(x)) if pd.notnull(x) and isinstance(x, float) and
                                x.is_integer() else str(x) if pd.notnull(x) else np.nan)

    # Load lookup table from destination file
    try:
        lookup_df = pd.read_excel(destination_file, sheet_name='ประเภทสินเชื่อ WM')
        lookup_df.columns = lookup_df.columns.str.strip()
        
        # Create lookup dictionary
        lookup_dict = dict(zip(lookup_df['คำอธิบาย'], lookup_df['รหัสประเภทสินเชื่อ WM']))
        
        # Map the values in zf50
        zf50['รหัสประเภทสินเชื่อ'] = zf50['ประเภทสินเชื่อ'].map(lookup_dict).fillna(zf50['รหัสประเภทสินเชื่อ'])
        
        print("✅ Successfully mapped รหัสประเภทสินเชื่อ using lookup table")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not perform lookup mapping - {e}")
        print("Using original รหัสประเภทสินเชื่อ values")

    zf50 = zf50.rename(columns={
    'รหัสประเภทสินเชื่อ': 'cont_group',
    'อัตราดอกเบี้ย': 'interest_flat_rate',
    'วันเริ่มต้น': 'first_due_date',
    'วันสิ้นสุด': 'last_due_date',
    'วันทำสัญญา': 'cont_date',
    'ลูกค้า': 'cust_code',
    'เลขที่จดทะเบียน VAT': 'ID',
    'หมายเลขถัง': 'chassis_no',
    'เงินต้น': 'principal',
    'จำนวนงวด': 'period',
    'ค่างวด': 'installment'
    })
    
    zf50['actual_date'] = zf50['cont_date']

    WM['เลขที่สัญญา'] = WM['เลขที่สัญญา'].apply(lambda x: str(int(x)) if pd.notnull(x) else np.nan)

    zf20_1['interest_accrued'] = pd.to_numeric(zf20_1['interest_accrued'].astype(str).str.replace(',', ''),
                                               errors='coerce').fillna(0.0)
    zf20_1['เลขที่สัญญา'] = zf20_1['เลขที่สัญญา'].apply(lambda x: str(int(x)) if pd.notnull(x) else np.nan)

    zf20_2['เลขที่สัญญา'] = zf20_2['เลขที่สัญญา'].apply(lambda x: str(int(x)) if pd.notnull(x) else np.nan)

    # convert to float
    columns_to_convert = ["installment_over_due","installment_current"]  
    
    for col in columns_to_convert:
        if col in zf20_2.columns:
            zf20_2[col] = zf20_2[col].astype(str).str.replace(",", "", regex=False)  # remove commas
            zf20_2[col] = pd.to_numeric(zf20_2[col], errors='coerce')

    # convert to float    
    if 'เลขที่สัญญา' in zf60.columns:
        zf60['เลขที่สัญญา'] = zf60['เลขที่สัญญา'].apply(lambda x: str(int(x)) if pd.notnull(x) else np.nan)
        
    def merge_column_with_same_ID(df1, df2, df3, df4, df5, id_column, how='left'):
        # Normalize column names
        for df in [df1, df2, df3, df4, df5]:
            df.columns = df.columns.str.strip().str.lower()
        id_column = id_column.lower()

        for i, df in enumerate([df1, df2, df3, df4, df5], start=1):
            if id_column not in df.columns:
                raise ValueError(f"DataFrame {i} does not contain '{id_column}'")

        # df2 aggregation: sum interest_accrued, take last next_due_date
        df2_sum_col = 'interest_accrued'
        df2_last_col = 'next_due_date'
        agg_funcs = {}
        if df2_sum_col in df2.columns:
            agg_funcs[df2_sum_col] = 'sum'
        if df2_last_col in df2.columns:
            agg_funcs[df2_last_col] = 'last'
        df2_agg = df2.groupby(id_column, as_index=False).agg(agg_funcs) if agg_funcs else df2.drop_duplicates(subset=[id_column])

        # df3 aggregation: sum all numeric columns except id_column
        df3_numeric_cols = [col for col in df3.columns if col != id_column and pd.api.types.is_numeric_dtype(df3[col])]
        df3_agg = df3.groupby(id_column, as_index=False)[df3_numeric_cols].sum() if df3_numeric_cols else df3.drop_duplicates(subset=[id_column])

        # Deduplicate df4 and df5
        df4_dedup = df4.drop_duplicates(subset=[id_column])
        df5_dedup = df5.drop_duplicates(subset=[id_column])

        # Merge step by step
        merged = df1.merge(df2_agg, on=id_column, how=how)
        merged = merged.merge(df3_agg, on=id_column, how=how)
        merged = merged.merge(df4_dedup, on=id_column, how=how)
        merged = merged.merge(df5_dedup, on=id_column, how=how)

        return merged

    merged_df = merge_column_with_same_ID(WM, zf20_1, zf20_2, zf60, zf50, id_column)

    merged_df['Ref. Aging'] = merged_df['เลขที่สัญญา']
    merged_df['send_bill_status'] =  "Y"
    merged_df['period_installment'] =  "1"

    source_file = merged_df

    # Change this to the actual name
    df_final = pd.read_excel(destination_file, sheet_name=target_sheet)

    # Step 1: Expand df_final to match the number of rows in source_file
    if df_final.empty:
        df_final = pd.DataFrame(columns=df_final.columns)  # keep column names
        df_final = df_final.reindex(index=range(len(source_file)))  # add blank rows
        
    column_mapping = {
    'เลขที่สัญญา': 'cont_no',
    'cont_date': 'cont_date', 
    'cust_code': 'cust_code',
    'id': 'ID',
    'chassis_no': 'chassis_no',
    'cont_group': 'cont_group',
    'send_bill_status': 'send_bill_status',
    'branch_code': 'branch_code',
    'actual_date': 'actual_date', 
    'principal': 'principal',
    'interest_flat_rate': 'interest_flat_rate',
    'period_installment': 'period_installment',
    'period': 'period',
    'installment': 'installment',
    'payment_day': 'payment_day',
    'first_due_date': 'first_due_date',
    'last_due_date': 'last_due_date',
    'current_balance': 'current_balance',
    'interest_accrued': 'interest_accrued',
    'next_due_date': 'next_due_date',
    'installment_current': 'installment_current',
    'installment_interest': 'installment_interest',
    'installment_over_due': 'installment_over_due',
    'dpd': 'dpd',
    'over_due_period': 'over_due_period',
    'Ref. Aging': 'Ref. Aging'
    }

    # Step 2: Copy over mapped data
    for src_col, dest_col in column_mapping.items():
        if src_col in source_file.columns and dest_col in df_final.columns:
            df_final[dest_col] = source_file[src_col].values  # row-wise copy
            
    #load temp9
    Template_9_df = pd.read_excel(Temp9_File)
    print("Load template 9 success")
        
    Aging_df = dowload_df(Aging_file)
    print("Load Aging success. Process check column...")
    
    filtered = Template_9_df[Template_9_df['payfor_code'] == 1001]
    
    # คำนวณผลรวม payment ของแต่ละ cont_no ใน filtered (จะได้ Series ที่ index เป็น cont_no)
    principal_paid = filtered.groupby('cont_no')['principal_paid'].sum()

    # จากนั้นเอาไปแมปใน Template_9_df หรือใช้ merge
    unique_template_9 = Template_9_df.drop_duplicates(subset=['cont_no']).copy()

    # แมปผลรวมลงใน unique_template_9
    unique_template_9['principal_paid_sum'] = unique_template_9['cont_no'].map(principal_paid)
    
    # Delete the "มูลหนี้คงเหลือ" column if it exists
    if 'มูลหนี้คงเหลือ' in df_final.columns:
        df_final = df_final.drop(columns=['มูลหนี้คงเหลือ'])
    
    df_final['มูลหนี้คงเหลือตาม_Aging'] = Aging_df[' ผลรวมรายการเปิ']

    # Ensure data types are compatible and handle NaN values for calculation
    # Convert both columns to numeric, handling any string values or NaN
    df_final['current_balance'] = pd.to_numeric(df_final['current_balance'], errors='coerce')
    df_final['มูลหนี้คงเหลือตาม_Aging'] = pd.to_numeric(df_final['มูลหนี้คงเหลือตาม_Aging'], errors='coerce')

    # Fill NaN values with 0 for calculation purposes
    current_balance_clean = df_final['current_balance'].fillna(0)
    aging_balance_clean = df_final['มูลหนี้คงเหลือตาม_Aging'].fillna(0)

    # Calculate the difference: current_balance - มูลหนี้คงเหลือตาม_Aging
    df_final['ตรวจ_diff_มูลหนี้คงเหลือ'] = current_balance_clean - aging_balance_clean

    print("✅ Successfully calculated ตรวจ_diff_มูลหนี้คงเหลือ with proper data type handling")
        
    if 'dpd' in df_final.columns:
        def calculate_overdue(dpd_value):
            try:
                dpd_float = float(dpd_value)
                period = math.floor(dpd_float / 30)
                return str(period) if period > 0 else ''
            except (ValueError, TypeError):
                return ''

    df_final['over_due_period'] = df_final['dpd'].apply(calculate_overdue)

    # --- Custom logic: Remove rows with blank cont_no and clear over_due_period if all 4 columns are blank ---
    # 1. Drop rows where 'cont_no' is blank or NaN
    if 'cont_no' in df_final.columns:
        df_final = df_final[df_final['cont_no'].notna() & (df_final['cont_no'].astype(str).str.strip() != '')]

    # 2. For rows where all 4 columns are blank/NaN, clear 'over_due_period'
    cols_to_check = ['installment_current', 'installment_interest', 'installment_over_due', 'dpd']
    if all(col in df_final.columns for col in cols_to_check + ['over_due_period']):
        mask_all_blank = df_final[cols_to_check].isna().all(axis=1) | (df_final[cols_to_check] == '').all(axis=1)
        df_final.loc[mask_all_blank, 'over_due_period'] = ''
    
    # --- Clear next_due_date if over_due_period is blank or NaN ---
    if 'next_due_date' in df_final.columns and 'over_due_period' in df_final.columns:
        mask_overdue_blank = df_final['over_due_period'].isna() | (df_final['over_due_period'].astype(str).str.strip() == '')
        df_final.loc[mask_overdue_blank, 'next_due_date'] = ''
    
    # 1. Copy 'ผลรวมรายการเปิ' from aging_df to df_final['มูลหนี้คงเหลือตาม_Aging']
    if ' ผลรวมรายการเปิ' in Aging_df.columns:
        df_final['มูลหนี้คงเหลือตาม_Aging'] = Aging_df[' ผลรวมรายการเปิ'].values
    else:
        print("❌ ' ผลรวมรายการเปิ' column not found in aging file.")

    # Ensure data types are compatible and handle NaN values for calculation
    # Convert both columns to numeric, handling any string values or NaN
    df_final['current_balance'] = pd.to_numeric(df_final['current_balance'], errors='coerce')
    df_final['มูลหนี้คงเหลือตาม_Aging'] = pd.to_numeric(df_final['มูลหนี้คงเหลือตาม_Aging'], errors='coerce')

    # Fill NaN values with 0 for calculation purposes
    current_balance_clean = df_final['current_balance'].fillna(0)
    aging_balance_clean = df_final['มูลหนี้คงเหลือตาม_Aging'].fillna(0)

    # Calculate the difference: current_balance - มูลหนี้คงเหลือตาม_Aging
    df_final['ตรวจ_diff_มูลหนี้คงเหลือ'] = current_balance_clean - aging_balance_clean

    print("✅ Successfully calculated ตรวจ_diff_มูลหนี้คงเหลือ with proper data type handling")

# 2. Compare with 'current_balance'
    if 'มูลหนี้คงเหลือตาม_Aging' in df_final.columns and 'current_balance' in df_final.columns:
        mismatch_mask = df_final['มูลหนี้คงเหลือตาม_Aging'] != df_final['current_balance']
        mismatches = df_final[mismatch_mask]
        
        if not mismatches.empty:
            print("⚠️ Mismatch found between 'มูลหนี้คงเหลือตาม_Aging' and 'current_balance':")
            print(mismatches[['มูลหนี้คงเหลือตาม_Aging', 'current_balance']])
        else:
            print("✅ All values match between 'มูลหนี้คงเหลือตาม_Aging' and 'current_balance'.")
    else:
        print("❌ Required columns not found for comparison.")

    # --- End custom logic ---

    # Create a Pandas ExcelWriter using the openpyxl engine
    if os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_final.to_excel(writer, sheet_name=target_sheet, index=False)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df_final.to_excel(writer, sheet_name=target_sheet, index=False)
                    
# source_file1 = r"D:\Angelway\Migration to python\File_testing\WM_test\Tem.6_WM\WM_1014100 - 1014108_04.2025.xls"
# zfloan_raw = r"D:\Angelway\Migration to python\File_testing\WM_test\Tem.6_WM\zfloan20_04.2025 ไฟล์ดิบ.txt"
# source_file3 = r"D:\Angelway\Migration to python\File_testing\WM_test\Tem.6_WM\zfloan50_04.2025.XLSX"
# source_file4 = r"D:\Angelway\Migration to python\File_testing\WM_test\Tem.6_WM\zfloan 60 04.2025.xlsx"
# Temp9_File = r"D:\Angelway\Migration to python\MigrationFunction\Angelway-Data_Migration_1\Template_9_WM_output.xlsx"
# Aging_File = r"D:\Angelway\Migration to python\File_testing\WL_test\Tem.7\Aging_04.2025.xls"
# destination_file = r"D:\Angelway\Migration to python\File_testing\WM_test\Tem.6_WM\6-ข้อมูลสัญญาเงินกู้ลดต้นลดดอก.xlsx"

# Migration_to_Template_6_WM(source_file1,zfloan_raw,source_file3,source_file4,Temp9_File,Aging_File,destination_file)
