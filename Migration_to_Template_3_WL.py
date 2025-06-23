from function.ColumnMappingFunction import map_excel_columns
from function.split_value_by_keyword import split_value_by_keyword
from function.replace_text_in_column import replace_text_in_column
import pandas as pd
from openpyxl import load_workbook

output_file = "Template_3_WL_output.xlsx"

def Migration_to_Template_3_WL(source_file1, source_sheet1, source_file2, source_sheet2, destination_file, destination_sheet):
    column_mapping = {
        'เลขตัวถังรถ': 'chassis_no',
    }

    # Map template columns
    template3_df = map_excel_columns(
        source_file2,
        destination_file,
        source_sheet2,
        destination_sheet,
        column_mapping
    )

    # Load headers from source_file1 using openpyxl
    wb = load_workbook(source_file1, data_only=True)
    ws = wb[source_sheet1]
    header_row_cells = ws[1]  # openpyxl is 1-based
    source_headers = [cell.value if cell.value is not None else f"Unnamed: {i}" 
                      for i, cell in enumerate(header_row_cells)]
    print("Headers from openpyxl:", source_headers)

    # Load source_file1 with pandas
    source_file1_df = pd.read_excel(source_file1, sheet_name=source_sheet1)
    source_file1_df.columns = source_headers
    # print(source_file1_df.head(5))

    # Columns to map from A to B
    columns_to_map = {
        'engine_no': 'engine_no',
        'product_type_code': 'product_type_code',
        'reg_no': 'reg_no',
        'brand_code': 'brand_code',
        'model_code': 'model_code_1',
        'sub_model_code': 'sub_model_code_1',
        'color_code': 'color_code_1',
        'rate_book': 'rate_book'
    }

    df_merged = template3_df.merge(
        source_file1_df[['chassis_no'] + list(columns_to_map.keys())],
        on='chassis_no',
        how='left',
        suffixes=('', '_from_a')
    )

    # Update columns in B using values from A
    for src_col, dest_col in columns_to_map.items():
        df_merged[dest_col] = df_merged[src_col + '_from_a']

    # Drop helper columns
    columns_to_drop = [col + '_from_a' for col in columns_to_map.keys()]
    template3_df = df_merged.drop(columns=columns_to_drop)

    # Post-processing steps
    steps = [
        ([
            "RED2","RED", "R-B", "BUS", "B-R", "B-S", "BUE", "R-S", "BLU", "W-R", "B-S_P", "R-G_P",
            "RED_P", "B-G", "BUB", "SLV", "S-B", "WBU", "W-B", "BLK", "R-B_P", "G-B_P",
            "2W-B", "SBW", "R-W", "BUR", "G-B", "S-R", "BRN", "P-W", "WHT", "BUW", "G-W",
            "GRN", "Y-W", "B-Y", "W-G", "Y-B", "RBU", "GNB", "GBU", "GRY", "BGO", "G-R",
            "O-B", "EBU", "BUG", "P-G", "B-P", "R-G", "YEL", "W-P", "W-Y", "B-W", "BRW",
            "WGO", "PKW", "RGO", "BUY", "R-S_P", "G-Y", "WPK", "W-S", "RBW", "O-W", "B-O",
            "BRS", "RBR", "BBR", "BBU", "S-Y"
        ], 'color_code'),
        ([
            "TH", "3TH", "4TH", "2TH", "TH2", "TH1", "TH3", "TH4",
            "TH6", "5TH", "9TH", "6TH", "8TH", "7TH",'(C)'
        ], 'sub_model_code')
    ]

    split_value_by_keyword(template3_df, 'sub_model_code_1', steps, 'model_code')

    replace_text_in_column(template3_df, 'sub_model_code', '(C)', 'C')
    replace_text_in_column(template3_df, 'color_code', 'RED2', 'RED')

    template3_df.fillna('-', inplace=True)
    template3_df.to_excel(output_file, index=False)
    print('🎉 Save success!')
